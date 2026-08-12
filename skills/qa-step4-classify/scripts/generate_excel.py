"""
qa-step4-classify 双页签 Excel 生成脚本

用法：
  python generate_excel.py <cases_md_path> <output_xlsx_path>

输入：04-cases.md（含 TABLE:cases，exec_by 已填定）
输出：04-cases.xlsx（双页签）

页签一「人类视图」：tc_id / module / biz_scene / title / precondition / test_content / expected / priority / exec_by
页签二「AI 视图」  ：人类视图九列 + entry_path / depends_on / test_data / assertions / tool_type / tool_payload / db_verify / cleanup / block_reason

依赖：openpyxl（pip install openpyxl）
"""

import re
import sys
import json

HUMAN_COLS = [
    "tc_id", "module", "biz_scene", "title",
    "precondition", "test_content", "expected",
    "priority", "exec_by"
]

AI_EXTRA_COLS = [
    "entry_path", "depends_on", "test_data", "assertions",
    "tool_type", "tool_payload", "db_verify", "cleanup", "block_reason"
]

# 表头中文映射
HEADER_CN = {
    "tc_id": "用例ID",
    "module": "所属模块",
    "biz_scene": "业务场景",
    "title": "用例标题",
    "precondition": "前置条件",
    "test_content": "测试内容",
    "expected": "预期结果",
    "priority": "优先级",
    "exec_by": "执行方",
    "entry_path": "入口路径",
    "depends_on": "依赖用例ID",
    "test_data": "测试数据",
    "assertions": "断言点",
    "tool_type": "工具类型",
    "tool_payload": "工具内容",
    "db_verify": "库表校验",
    "cleanup": "清理动作",
    "block_reason": "阻塞原因",
}


def parse_markdown_table(md_path):
    """从 Markdown 文件解析 TABLE:cases"""
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()

    # 找锚点
    table_match = re.search(
        r"<!-- TABLE:cases BEGIN -->(.*?)<!-- TABLE:cases END -->",
        content, re.DOTALL
    )
    if not table_match:
        # 注：fallback 分支必须让 group(1) 也包含表头行，否则表头被误当首条数据，
        # 导致 header_line 错位、tc_id 取不到、整片解析为 0 行。与 merge_cls.py 保持一致。
        table_match = re.search(
            r"(\| tc_id \|.*?\n\|[-| ]+\|\n(?:\|.*\|\n)+)",
            content
        )

    if not table_match:
        return [], "未找到 TABLE:cases"

    table_text = table_match.group(1) if table_match.lastindex == 1 else table_match.group(1)
    lines = table_text.strip().split("\n")
    if len(lines) < 2:
        return [], "表格为空"

    # 解析表头
    header_line = lines[0]
    headers = [h.strip() for h in header_line.strip("|").split("|")]
    data_start = 1
    cases = []

    for line in lines[data_start:]:
        if not line.strip().startswith("|"):
            continue
        # 跳过分隔线（如 |------|------|）
        _stripped = line.strip().strip("|")
        if set(_stripped.replace("|", "").replace(" ", "")) <= set("-:"):
            continue
        cells = [c.strip() for c in _stripped.split("|")]
        if len(cells) < 5:
            continue
        case = {}
        for i, h in enumerate(headers):
            if i < len(cells):
                case[h] = cells[i]
        if case.get("tc_id"):
            cases.append(case)

    return cases, None


def generate_excel(cases, output_path):
    """生成双页签 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("❌ 缺少 openpyxl，请安装: pip install openpyxl")
        return False

    wb = Workbook()

    # ---- 样式 ----
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ai_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    human_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # ---- 页签一：人类视图 ----
    ws1 = wb.active
    ws1.title = "人类视图"

    # 写表头
    for col_idx, col_name in enumerate(HUMAN_COLS, 1):
        cell = ws1.cell(row=1, column=col_idx, value=HEADER_CN.get(col_name, col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据
    for row_idx, case in enumerate(cases, 2):
        for col_idx, col_name in enumerate(HUMAN_COLS, 1):
            value = case.get(col_name, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if case.get("exec_by") == "AI":
                cell.fill = ai_fill
            elif case.get("exec_by") == "人":
                cell.fill = human_fill

    # 设置列宽
    col_widths_human = {
        "tc_id": 10, "module": 12, "biz_scene": 18, "title": 25,
        "precondition": 20, "test_content": 25, "expected": 30,
        "priority": 8, "exec_by": 8
    }
    for col_idx, col_name in enumerate(HUMAN_COLS, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = col_widths_human.get(col_name, 15)

    # ---- 页签二：AI 视图 ----
    ws2 = wb.create_sheet(title="AI 视图")

    # 写表头（人类列 + AI 追加列）
    all_cols = HUMAN_COLS + AI_EXTRA_COLS
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=HEADER_CN.get(col_name, col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据
    for row_idx, case in enumerate(cases, 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            value = case.get(col_name, "")
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 设置列宽
    for col_idx, col_name in enumerate(all_cols, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 18

    # ---- 冻结首行 ----
    ws1.freeze_panes = "A2"
    ws2.freeze_panes = "A2"

    # ---- 自动筛选 ----
    ws1.auto_filter.ref = ws1.dimensions
    ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)
    return True


def main():
    if len(sys.argv) < 3:
        print("用法: python generate_excel.py <cases_md_path> <output_xlsx_path> [--json]")
        sys.exit(1)

    md_path = sys.argv[1]
    output_path = sys.argv[2]
    output_json = "--json" in sys.argv

    cases, error = parse_markdown_table(md_path)
    if error:
        result = {"error": error, "success": False}
        if output_json:
            print(json.dumps(result, ensure_ascii=False))
        else:
            print(f"❌ {error}")
        sys.exit(1)

    if not cases:
        result = {"error": "无用例数据", "success": False}
        if output_json:
            print(json.dumps(result, ensure_ascii=False))
        else:
            print("❌ 无用例数据")
        sys.exit(1)

    success = generate_excel(cases, output_path)

    # 统计
    ai_count = sum(1 for c in cases if c.get("exec_by") == "AI")
    human_count = sum(1 for c in cases if c.get("exec_by") == "人")
    pending_count = sum(1 for c in cases if c.get("exec_by") not in ("AI", "人"))

    result = {
        "success": success,
        "output": output_path,
        "total": len(cases),
        "ai_count": ai_count,
        "human_count": human_count,
        "pending_count": pending_count,
    }

    if output_json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f"✅ Excel 已生成: {output_path}")
        print(f"   总用例: {len(cases)}  |  AI 执行: {ai_count}  |  人执行: {human_count}  |  待分类: {pending_count}")


if __name__ == "__main__":
    main()
