"""
qa-step3-cases 用例格式化导出工具

用法：
  python format_excel.py <cases_md_path> <output_xlsx_path>

功能：把 Markdown 中的 TABLE:cases 原样导出为格式化 Excel（单页签，人类视图九列）。
不做任何校验——校验是 AI agent 的活，不是代码的活。
这只是个"把表格从 md 搬到 xlsx"的机械工具。

输出列：tc_id / module / biz_scene / title / precondition / test_content / expected / priority / exec_by

依赖：openpyxl
"""

import re
import sys
import json

COLS = [
    "tc_id", "module", "biz_scene", "title",
    "precondition", "test_content", "expected",
    "priority", "exec_by"
]

HEADER_CN = {
    "tc_id": "用例ID", "module": "所属模块", "biz_scene": "业务场景",
    "title": "用例标题", "precondition": "前置条件", "test_content": "测试内容",
    "expected": "预期结果", "priority": "优先级", "exec_by": "执行方",
}


def parse_markdown_table(md_path):
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()

    match = re.search(
        r"<!-- TABLE:cases BEGIN -->(.*?)<!-- TABLE:cases END -->",
        content, re.DOTALL
    )
    if not match:
        match = re.search(
            r"\| tc_id \|.*?\n\|[-| ]+\|\n((?:\|.*\|\n)+)",
            content
        )
    if not match:
        return []

    table_text = match.group(1) if match.lastindex == 1 else match.group(1)
    lines = [l for l in table_text.strip().split("\n") if l.strip().startswith("|")]
    if len(lines) < 2:
        return []

    headers = [h.strip() for h in lines[0].strip("|").split("|")]
    cases = []
    for line in lines[1:]:
        cells = [c.strip() for c in line.strip("|").split("|")]
        if len(cells) < 5:
            continue
        case = {headers[i]: cells[i] for i in range(min(len(headers), len(cells)))}
        if case.get("tc_id"):
            cases.append(case)
    return cases


def write_excel(cases, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "用例"

    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=HEADER_CN.get(col, col))
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border

    for ri, case in enumerate(cases, 2):
        for ci, col in enumerate(COLS, 1):
            c = ws.cell(row=ri, column=ci, value=case.get(col, ""))
            c.alignment = cell_align; c.border = border

    widths = {"tc_id": 10, "module": 12, "biz_scene": 18, "title": 25,
              "precondition": 20, "test_content": 28, "expected": 30, "priority": 8, "exec_by": 8}
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = widths.get(col, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def main():
    if len(sys.argv) < 3:
        print("用法: python format_excel.py <cases_md> <output_xlsx> [--json]", file=sys.stderr)
        sys.exit(1)

    md, out = sys.argv[1], sys.argv[2]
    cases = parse_markdown_table(md)
    if not cases:
        print("❌ 未找到用例数据")
        sys.exit(1)

    write_excel(cases, out)
    info = {"output": out, "count": len(cases)}
    if "--json" in sys.argv:
        print(json.dumps(info, ensure_ascii=False))
    else:
        print(f"✅ {out}  ({len(cases)} 条用例)")


if __name__ == "__main__":
    main()
