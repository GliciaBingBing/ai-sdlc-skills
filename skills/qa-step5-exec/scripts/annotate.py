"""
qa-step5-exec 执行结果标注脚本

用法：
  python annotate.py <04-cases.xlsx> <05-results.json> <05-results.xlsx>

功能：读原始用例 Excel（AI 视图页签），把执行结果逐行写进去，不做任何汇总。

输出：05-results.xlsx（AI 视图页签 + 执行结果列，原格式不动）
- 在 AI 视图页签右侧追加四列：执行状态 / 失败断言 / 实际结果 / 证据路径
- 其他页签（人类视图）原样保留
- 不做覆盖率统计、不做摘要、不做 Markdown 报告

依赖：openpyxl
"""

import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


STATUS_FILLS = {
    "通过": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "失败": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "阻塞": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "未执行-红线待确认": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "未执行-需人工": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "未执行-依赖未通过": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

RESULT_HEADERS = ["执行状态", "失败断言", "实际结果", "证据路径"]


def main():
    if len(sys.argv) < 4:
        print("用法: python annotate.py <04-cases.xlsx> <05-results.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    json_path = sys.argv[2]
    output_path = sys.argv[3]

    # 读结果
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    results = {r["tc_id"]: r for r in data.get("results", [])}

    # 读 Excel
    wb = load_workbook(xlsx_path)

    # 找 AI 视图页签：优先精确名「AI 视图」，其次含「AI」，最后含「视图」（避免误命中「人类视图」）
    ai_sheet = None
    for name in wb.sheetnames:
        if name == "AI 视图":
            ai_sheet = wb[name]
            break
    if ai_sheet is None:
        for name in wb.sheetnames:
            if "AI" in name:
                ai_sheet = wb[name]
                break
    if ai_sheet is None:
        for name in wb.sheetnames:
            if "视图" in name:
                ai_sheet = wb[name]
                break
    if ai_sheet is None:
        ai_sheet = wb.active

    # 找 tc_id 列（兼容中文规范表头「用例ID」与英文「tc_id」）
    tc_col = None
    for col in range(1, ai_sheet.max_column + 1):
        v = str(ai_sheet.cell(row=1, column=col).value or "").strip()
        if v in ("用例ID", "tc_id", "用例ID "):
            tc_col = col
            break
    if tc_col is None:
        print("❌ 未找到「用例ID」列（表头需含 用例ID 或 tc_id）")
        sys.exit(1)

    # 在末尾追加四列表头
    hdr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    hdr_font = Font(bold=True)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    start_col = ai_sheet.max_column + 1
    for i, h in enumerate(RESULT_HEADERS):
        c = ai_sheet.cell(row=1, column=start_col + i, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin_border

    # 逐行标注
    annotated = 0
    for row in range(2, ai_sheet.max_row + 1):
        tc_id = str(ai_sheet.cell(row=row, column=tc_col).value or "").strip()
        if not tc_id:
            continue

        result = results.get(tc_id)
        if not result:
            continue

        status = result.get("status", "")
        values = [
            status,
            result.get("failed_assertion", ""),
            result.get("actual", ""),
            result.get("evidence_path", ""),
        ]

        for i, v in enumerate(values):
            c = ai_sheet.cell(row=row, column=start_col + i, value=v)
            c.alignment = cell_align; c.border = thin_border
            if i == 0 and status in STATUS_FILLS:
                c.fill = STATUS_FILLS[status]

        annotated += 1

    # 列宽
    for i in range(4):
        ai_sheet.column_dimensions[ai_sheet.cell(row=1, column=start_col + i).column_letter].width = 18

    wb.save(output_path)

    not_found = len(results) - annotated
    print(f"✅ 已标注 {annotated} 条结果 → {output_path}")
    if not_found > 0:
        print(f"   ⚠️ {not_found} 条结果在 Excel 中未找到对应用例ID")


if __name__ == "__main__":
    main()
